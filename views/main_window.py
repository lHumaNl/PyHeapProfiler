import logging
import csv
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from PySide6.QtWidgets import (
    QMainWindow, QVBoxLayout, QWidget, QPushButton,
    QTableView, QLabel, QHeaderView, QMessageBox, QMenu, QProgressBar, QMenuBar, QToolBar,
    QLineEdit, QComboBox, QCheckBox, QHBoxLayout, QGroupBox, QSplitter
)
from PySide6.QtGui import QStandardItemModel, QStandardItem, QAction
from PySide6.QtCore import Qt, QSortFilterProxyModel, QObject, Signal


class RowNumberProxyModel(QSortFilterProxyModel):
    """Proxy model that displays visual row numbers in vertical header."""

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if orientation == Qt.Orientation.Vertical and role == Qt.ItemDataRole.DisplayRole:
            # Return visual row number (1-based) for vertical header
            return str(section + 1)
        return super().headerData(section, orientation, role)
from views.object_details_window import ObjectDetailsWindow
from controllers.main_controller import MainController
from utils.helpers import int_or_none, float_or_none
from utils.settings_manager import SettingsManager
from utils.theme_manager import ThemeManager
from utils.error_handler import show_error_message, get_traceback


class MainWindow(QMainWindow):
    """
    The main window of the application.
    """

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Heap Dump Analyzer")
        self.resize(800, 600)
        self.logger = logging.getLogger(__name__)

        self.current_dump = None
        self.comparison_dump = None
        self.details_windows = []
        self.chart_windows = []
        self.loading_worker = None
        self.loading_thread = None

        self.controller = MainController(self)
        self.settings_manager = SettingsManager()
        self.theme_manager = ThemeManager(self)

        self.current_theme = self.settings_manager.load_theme()

        self.setAcceptDrops(True)
        self.setup_ui()
        self.setup_menu()
        self.setup_toolbar()
        self.load_settings()

    def setup_ui(self):
        """
        Set up the user interface components.
        """
        self.main_widget = QWidget()
        self.setCentralWidget(self.main_widget)

        self.main_layout = QVBoxLayout()
        self.main_widget.setLayout(self.main_layout)

        # Load and control buttons
        buttons_layout = QHBoxLayout()
        self.load_button = QPushButton("Load Heap Dump\n(JSON)")
        self.compare_button = QPushButton("Load for\nComparison")
        self.compare_button.setEnabled(False)
        self.chart_button = QPushButton("Show Charts")
        self.chart_button.setEnabled(False)
        self.export_button = QPushButton("Export Results")
        self.export_button.setEnabled(False)

        buttons_layout.addWidget(self.load_button)
        buttons_layout.addWidget(self.compare_button)
        buttons_layout.addWidget(self.chart_button)
        buttons_layout.addWidget(self.export_button)
        self.main_layout.addLayout(buttons_layout)

        # Filter and search panel
        self.setup_filter_search_panel()

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setRange(0, 100)
        self.main_layout.addWidget(self.progress_bar)

        self.table_view = QTableView()
        self.main_layout.addWidget(self.table_view)

        self.summary_label = QLabel("Summary data will be displayed here.")
        self.main_layout.addWidget(self.summary_label)

        self.load_button.clicked.connect(self.controller.handle_load_dump)
        self.compare_button.clicked.connect(self.controller.handle_compare_dumps)
        self.table_view.clicked.connect(self.controller.handle_table_click)
        self.chart_button.clicked.connect(self.controller.handle_show_charts)
        self.export_button.clicked.connect(self.controller.handle_export)
        self.table_view.horizontalHeader().sectionMoved.connect(self.save_column_settings)
        self.table_view.horizontalHeader().sectionResized.connect(self.save_column_settings)
        self.table_view.horizontalHeader().sortIndicatorChanged.connect(self.save_sorting_settings)

        self.table_view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table_view.customContextMenuRequested.connect(self.show_column_menu)

    def setup_filter_search_panel(self):
        """
        Set up the filter and search panel.
        """
        filter_search_layout = QHBoxLayout()
        
        # Filter group
        filter_group = QGroupBox("Filters")
        filter_layout = QVBoxLayout()
        
        # Size filter
        size_filter_layout = QHBoxLayout()
        size_filter_layout.addWidget(QLabel("Min Size:"))
        self.min_size_input = QLineEdit()
        self.min_size_input.setPlaceholderText("bytes")
        self.min_size_input.setMaximumWidth(100)
        size_filter_layout.addWidget(self.min_size_input)
        size_filter_layout.addWidget(QLabel("Max Size:"))
        self.max_size_input = QLineEdit()
        self.max_size_input.setPlaceholderText("bytes")
        self.max_size_input.setMaximumWidth(100)
        size_filter_layout.addWidget(self.max_size_input)
        filter_layout.addLayout(size_filter_layout)
        
        # Object type filter
        type_filter_layout = QHBoxLayout()
        type_filter_layout.addWidget(QLabel("Object Type:"))
        self.type_filter_combo = QComboBox()
        self.type_filter_combo.setMinimumWidth(200)
        type_filter_layout.addWidget(self.type_filter_combo)
        filter_layout.addLayout(type_filter_layout)
        
        # Status filter (for comparison mode)
        self.status_filter_layout = QHBoxLayout()
        self.status_filter_layout.addWidget(QLabel("Status:"))
        self.status_checkboxes = {}
        for status in ['New', 'Deleted', 'Old', 'Modified']:
            checkbox = QCheckBox(status)
            self.status_checkboxes[status] = checkbox
            self.status_filter_layout.addWidget(checkbox)
        filter_layout.addLayout(self.status_filter_layout)
        
        filter_group.setLayout(filter_layout)
        filter_search_layout.addWidget(filter_group)
        
        # Search group
        search_group = QGroupBox("Search")
        search_layout = QVBoxLayout()
        
        # Object ID search
        id_search_layout = QHBoxLayout()
        id_search_layout.addWidget(QLabel("Object ID:"))
        self.search_id_input = QLineEdit()
        self.search_id_input.setPlaceholderText("Enter object ID...")
        id_search_layout.addWidget(self.search_id_input)
        search_layout.addLayout(id_search_layout)
        
        # Attribute value search
        attr_search_layout = QHBoxLayout()
        attr_search_layout.addWidget(QLabel("Attribute Value:"))
        self.search_attr_input = QLineEdit()
        self.search_attr_input.setPlaceholderText("Enter attribute value...")
        attr_search_layout.addWidget(self.search_attr_input)
        search_layout.addLayout(attr_search_layout)
        
        search_group.setLayout(search_layout)
        filter_search_layout.addWidget(search_group)
        
        # Apply filters button
        self.apply_filters_button = QPushButton("Apply Filters")
        self.apply_filters_button.clicked.connect(self.apply_filters)
        filter_search_layout.addWidget(self.apply_filters_button)
        
        # Reset filters button
        self.reset_filters_button = QPushButton("Reset Filters")
        self.reset_filters_button.clicked.connect(self.reset_filters)
        filter_search_layout.addWidget(self.reset_filters_button)
        
        self.main_layout.addLayout(filter_search_layout)
        
        # Hide status filter initially
        self.set_status_filter_visible(False)

    def set_status_filter_visible(self, visible):
        """
        Show/hide status filter based on comparison mode.

        Args:
            visible (bool): Whether to show the status filter.
        """
        for i in range(self.status_filter_layout.count()):
            widget = self.status_filter_layout.itemAt(i).widget()
            if widget:
                widget.setVisible(visible)

    def on_header_clicked(self, logical_index):
        """
        Handle header click events.

        Args:
            logical_index (int): The index of the clicked header section.
        """
        pass  # All columns are now sortable

    def update_progress(self, current, total):
        """
        Update the progress bar during loading.

        Args:
            current: Current bytes read.
            total: Total bytes to read.
        """
        if total > 0:
            progress = int((current / total) * 100)
            self.progress_bar.setValue(progress)

    def show_loading(self):
        """Show progress bar and disable buttons during loading."""
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)
        self.load_button.setEnabled(False)
        self.compare_button.setEnabled(False)

    def hide_loading(self):
        """Hide progress bar and enable buttons after loading."""
        self.progress_bar.setVisible(False)
        self.load_button.setEnabled(True)
        if self.current_dump:
            self.compare_button.setEnabled(True)
            self.chart_button.setEnabled(True)
            self.export_button.setEnabled(True)

    def setup_menu(self):
        """Set up the menu bar."""
        menubar = self.menuBar()

        file_menu = menubar.addMenu("File")

        load_action = QAction("Load Heap Dump", self)
        load_action.setShortcut("Ctrl+O")
        load_action.triggered.connect(self.controller.handle_load_dump)
        file_menu.addAction(load_action)

        compare_action = QAction("Load for Comparison", self)
        compare_action.setShortcut("Ctrl+C")
        compare_action.triggered.connect(self.controller.handle_compare_dumps)
        compare_action.setEnabled(False)
        self.compare_menu_action = compare_action
        file_menu.addAction(compare_action)

        export_action = QAction("Export Results", self)
        export_action.setShortcut("Ctrl+E")
        export_action.triggered.connect(self.controller.handle_export)
        export_action.setEnabled(False)
        self.export_menu_action = export_action
        file_menu.addAction(export_action)

        file_menu.addSeparator()

        refresh_action = QAction("Refresh", self)
        refresh_action.setShortcut("F5")
        refresh_action.triggered.connect(self.controller.handle_refresh)
        file_menu.addAction(refresh_action)

        file_menu.addSeparator()

        exit_action = QAction("Exit", self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        view_menu = menubar.addMenu("View")

        charts_action = QAction("Memory Charts", self)
        charts_action.triggered.connect(self.controller.handle_show_charts)
        charts_action.setEnabled(False)
        self.charts_menu_action = charts_action
        view_menu.addAction(charts_action)

        view_menu.addSeparator()

        theme_menu = view_menu.addMenu("Theme")

        light_theme_action = QAction("Light Theme", self)
        light_theme_action.setCheckable(True)
        light_theme_action.setChecked(self.current_theme == ThemeManager.LIGHT_THEME)
        light_theme_action.triggered.connect(lambda: self.set_theme(ThemeManager.LIGHT_THEME))
        self.light_theme_action = light_theme_action
        theme_menu.addAction(light_theme_action)

        dark_theme_action = QAction("Dark Theme", self)
        dark_theme_action.setCheckable(True)
        dark_theme_action.setChecked(self.current_theme == ThemeManager.DARK_THEME)
        dark_theme_action.triggered.connect(lambda: self.set_theme(ThemeManager.DARK_THEME))
        self.dark_theme_action = dark_theme_action
        theme_menu.addAction(dark_theme_action)

    def setup_toolbar(self):
        """Set up the toolbar."""
        toolbar = QToolBar("Main Toolbar")
        self.addToolBar(toolbar)

        load_action = QAction("Load", self)
        load_action.setStatusTip("Load a heap dump file")
        load_action.triggered.connect(self.controller.handle_load_dump)
        toolbar.addAction(load_action)

        compare_action = QAction("Compare", self)
        compare_action.setStatusTip("Load a heap dump for comparison")
        compare_action.setEnabled(False)
        compare_action.triggered.connect(self.controller.handle_compare_dumps)
        self.compare_toolbar_action = compare_action
        toolbar.addAction(compare_action)

        export_action = QAction("Export", self)
        export_action.setStatusTip("Export results to file")
        export_action.setEnabled(False)
        export_action.triggered.connect(self.controller.handle_export)
        self.export_toolbar_action = export_action
        toolbar.addAction(export_action)

        charts_action = QAction("Charts", self)
        charts_action.setStatusTip("Show memory usage charts")
        charts_action.setEnabled(False)
        charts_action.triggered.connect(self.controller.handle_show_charts)
        self.charts_toolbar_action = charts_action
        toolbar.addAction(charts_action)

        toolbar.addSeparator()

        theme_toggle_action = QAction("Toggle Theme", self)
        theme_toggle_action.setStatusTip("Switch between light and dark theme")
        theme_toggle_action.triggered.connect(self.toggle_theme)
        toolbar.addAction(theme_toggle_action)

    def set_theme(self, theme_name):
        """Set the application theme."""
        self.current_theme = theme_name
        ThemeManager.apply_theme(theme_name)
        self.settings_manager.save_theme(theme_name)

        self.light_theme_action.setChecked(theme_name == ThemeManager.LIGHT_THEME)
        self.dark_theme_action.setChecked(theme_name == ThemeManager.DARK_THEME)

        ThemeManager.apply_menu_theme(self.menuBar())

        self.logger.info(f"Theme changed to {theme_name}")

    def toggle_theme(self):
        """Toggle between light and dark themes."""
        new_theme = ThemeManager.DARK_THEME if self.current_theme == ThemeManager.LIGHT_THEME else ThemeManager.LIGHT_THEME
        self.set_theme(new_theme)

    def dragEnterEvent(self, event):
        """Handle drag enter event for file dropping."""
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls and urls[0].isLocalFile():
                file_path = urls[0].toLocalFile()
                if file_path.endswith('.json'):
                    event.acceptProposedAction()
                    self.logger.debug(f"Accepting drag of file: {file_path}")

    def dragMoveEvent(self, event):
        """Handle drag move event."""
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        """Handle drop event for loading heap dump files."""
        urls = event.mimeData().urls()
        if urls and urls[0].isLocalFile():
            file_path = urls[0].toLocalFile()
            if file_path.endswith('.json'):
                self.logger.info(f"File dropped: {file_path}")
                self.load_dump_from_path(file_path)

    def load_dump_from_path(self, file_path):
        """Load a heap dump from the given file path."""
        try:
            from models.heap_dump import HeapDumpModel
            self.current_dump = HeapDumpModel(file_path)
            self.current_dump.load_data()
            self.populate_table(self.current_dump)
            self.compare_button.setEnabled(True)
            self.chart_button.setEnabled(True)
            self.export_button.setEnabled(True)

            if hasattr(self, 'compare_toolbar_action'):
                self.compare_toolbar_action.setEnabled(True)
            if hasattr(self, 'compare_menu_action'):
                self.compare_menu_action.setEnabled(True)
            if hasattr(self, 'export_toolbar_action'):
                self.export_toolbar_action.setEnabled(True)
            if hasattr(self, 'export_menu_action'):
                self.export_menu_action.setEnabled(True)
            if hasattr(self, 'charts_toolbar_action'):
                self.charts_toolbar_action.setEnabled(True)
            if hasattr(self, 'charts_menu_action'):
                self.charts_menu_action.setEnabled(True)

            self.show_message("Success", f"Heap dump loaded successfully\n{file_path}")
            self.logger.info(f"Heap dump loaded from {file_path}")
        except Exception as e:
            self.logger.exception("Failed to load heap dump")
            show_error_message(self, "Loading Error", f"Failed to load heap dump: {str(e)}", get_traceback())

    def load_settings(self):
        """Load application settings."""
        self.settings_manager.load_window_geometry(self)

        column_visibility = self.settings_manager.load_column_visibility()
        column_order = self.settings_manager.load_column_order()
        column_widths = self.settings_manager.load_column_widths()
        sort_column, sort_order = self.settings_manager.load_sorting()

        self.column_visibility = column_visibility
        self.column_order = column_order
        self.column_widths = column_widths
        self.sort_column = sort_column
        self.sort_order = sort_order

        self.set_theme(self.current_theme)

        self.logger.info("Settings loaded successfully")

    def save_settings(self):
        """Save application settings."""
        self.settings_manager.save_window_geometry(self)
        self.logger.info("Settings saved successfully")

    def save_column_settings(self):
        """Save column-related settings."""
        if not hasattr(self, 'table_view') or not self.table_view.model():
            return

        model = self.table_view.model()
        hidden_columns = []
        column_order = []
        column_widths = []

        for col in range(model.columnCount()):
            if self.table_view.isColumnHidden(col):
                hidden_columns.append(col)
            column_order.append(self.table_view.horizontalHeader().visualIndex(col))
            column_widths.append(self.table_view.columnWidth(col))

        self.settings_manager.save_column_visibility(hidden_columns)
        self.settings_manager.save_column_order(column_order)
        self.settings_manager.save_column_widths(column_widths)

    def save_sorting_settings(self, column, order):
        """Save sorting settings."""
        self.settings_manager.save_sorting(column, order)

    def apply_column_settings(self):
        """Apply saved column settings to the table."""
        if not hasattr(self, 'table_view') or not self.table_view.model():
            return

        model = self.table_view.model()

        if hasattr(self, 'column_visibility') and self.column_visibility:
            for col in self.column_visibility:
                if col < model.columnCount():
                    self.table_view.setColumnHidden(col, True)

        if hasattr(self, 'column_order') and self.column_order:
            header = self.table_view.horizontalHeader()
            for logical_col, visual_pos in enumerate(self.column_order):
                if visual_pos < model.columnCount():
                    header.moveSection(header.visualIndex(logical_col), visual_pos)

        if hasattr(self, 'column_widths') and self.column_widths:
            for col, width in enumerate(self.column_widths):
                if col < model.columnCount():
                    self.table_view.setColumnWidth(col, width)

        if hasattr(self, 'sort_column') and hasattr(self, 'sort_order'):
            if self.sort_column >= 0 and self.sort_column < model.columnCount():
                self.table_view.sortByColumn(self.sort_column, self.sort_order)

    def closeEvent(self, event):
        """Handle window close event."""
        self.save_settings()
        self.logger.info("Application closing")
        super().closeEvent(event)

    def populate_table(self, data, comparison=False):
        """
        Populate the main table with data.

        Args:
            data (HeapDumpModel): The heap dump data to display.
            comparison (bool): Whether to include comparison data.
        """
        try:
            if not data:
                return

            model = QStandardItemModel()
            headers = [
                "Object\nType", "Number of\nObjects", "% of\nTotal\nObjects",
                "Total Size\n(bytes)", "% of\nTotal\nSize"
            ]

            if comparison:
                headers.extend([
                    "Objects\n(Main)", "Objects\n(Comparison)",
                    "New\nObjects", "Deleted\nObjects",
                    "Size\n(Main)", "Size\n(Comparison)",
                    "Size\nChange\n(%)", "Size\nChange\n(bytes)"
                ])

            model.setHorizontalHeaderLabels(headers)

            total_objects = data.total_objects if hasattr(data, 'total_objects') else 0
            total_size = data.total_size if hasattr(data, 'total_size') else 0

            comparison_results = None
            if comparison and self.current_dump and self.comparison_dump:
                comparison_results = self.current_dump.compare_with(self.comparison_dump)
                all_obj_types = set(comparison_results.keys())
            else:
                all_obj_types = data.processed_data.keys() if hasattr(data, 'processed_data') else []

            # Update type filter combo with all available types
            if self.current_dump and hasattr(self.current_dump, 'processed_data'):
                all_types = list(self.current_dump.processed_data.keys())
                self.update_type_filter_combo(all_types)

            for obj_type in all_obj_types:
                type_item = QStandardItem(obj_type)
                type_item.setEditable(False)
                type_item.setForeground(Qt.GlobalColor.blue)
                type_item.setData(obj_type, Qt.ItemDataRole.UserRole)

                if comparison and self.current_dump and self.comparison_dump and comparison_results:
                    comp_result = comparison_results.get(obj_type, {})
                    num_objects_main = comp_result.get('num_objects_main', 0)
                    num_objects_other = comp_result.get('num_objects_other', 0)
                    num_new_objects = comp_result.get('num_new_objects', 0)
                    num_deleted_objects = comp_result.get('num_deleted_objects', 0)

                    total_size_main = comp_result.get('total_size_main', 0)
                    total_size_other = comp_result.get('total_size_other', 0)
                    size_change = comp_result.get('size_change', 0)
                    size_percent_change = comp_result.get('size_percent_change', 0)

                    num_total_objects = num_objects_main + num_new_objects - num_deleted_objects
                    total_objects_combined = self.current_dump.total_objects + self.comparison_dump.total_objects
                    perc_objects = (num_total_objects / total_objects_combined * 100) if total_objects_combined else 0

                    total_size_combined = total_size_main + size_change
                    total_size_overall = self.current_dump.total_size + self.comparison_dump.total_size
                    perc_size = (total_size_combined / total_size_overall * 100) if total_size_overall else 0

                    row_items = [
                        type_item,
                        QStandardItem(int_or_none(num_total_objects)),
                        QStandardItem(float_or_none(perc_objects)),
                        QStandardItem(int_or_none(total_size_combined)),
                        QStandardItem(float_or_none(perc_size)),
                        QStandardItem(int_or_none(num_objects_main)),
                        QStandardItem(int_or_none(num_objects_other)),
                        QStandardItem(int_or_none(num_new_objects)),
                        QStandardItem(int_or_none(num_deleted_objects)),
                        QStandardItem(int_or_none(total_size_main)),
                        QStandardItem(int_or_none(total_size_other)),
                        QStandardItem(float_or_none(size_percent_change)),
                        QStandardItem(int_or_none(size_change))
                    ]
                else:
                    obj_data = data.processed_data[obj_type]
                    num_objects = obj_data['num_objects']
                    total_size_obj = obj_data['total_size']

                    perc_objects = (num_objects / total_objects * 100) if total_objects else 0
                    perc_size = (total_size_obj / total_size * 100) if total_size else 0

                    row_items = [
                        type_item,
                        QStandardItem(int_or_none(num_objects)),
                        QStandardItem(float_or_none(perc_objects)),
                        QStandardItem(int_or_none(total_size_obj)),
                        QStandardItem(float_or_none(perc_size))
                    ]

                for item in row_items[1:]:
                    try:
                        value = float(item.text()) if '.' in item.text() else int(item.text())
                        item.setData(value, Qt.ItemDataRole.UserRole)
                    except (ValueError, AttributeError):
                        pass

                model.appendRow(row_items)

            # Set up proxy model for numeric sorting with dynamic row numbers
            self.proxy_model = RowNumberProxyModel()
            self.proxy_model.setSourceModel(model)
            self.proxy_model.setSortRole(Qt.ItemDataRole.UserRole)
            self.table_view.setModel(self.proxy_model)

            self.table_view.setSortingEnabled(True)
            self.table_view.sortByColumn(1, Qt.SortOrder.DescendingOrder)

            self.table_view.horizontalHeader().sectionClicked.connect(self.on_header_clicked)
            header = self.table_view.horizontalHeader()
            header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
            self.table_view.resizeColumnsToContents()
            self.table_view.horizontalHeader().setStretchLastSection(False)

            # Adjust window size to fit columns
            total_width = self.table_view.verticalHeader().width()
            for i in range(model.columnCount()):
                total_width += self.table_view.columnWidth(i)
            total_width += self.table_view.verticalScrollBar().sizeHint().width()
            self.resize(total_width + 50, self.height())

            # Show/hide status filter based on comparison mode
            self.set_status_filter_visible(comparison)

            self.summary_label.setText(
                f"Total objects: {total_objects}, Total size: {total_size} bytes"
            )

            self.logger.info("Main table populated successfully")
        except Exception as e:
            self.logger.exception("Failed to populate main table")
            show_error_message(self, "Table Error", f"Failed to populate table: {str(e)}", get_traceback())

    def show_message(self, title, message):
        """
        Display a message box.

        Args:
            title (str): The title of the message box.
            message (str): The message to display.
        """
        QMessageBox.information(self, title, message)

    def open_object_details(self, obj_type):
        """
        Open the object details window for a specific object type.

        Args:
            obj_type (str): The object type to display.
        """
        try:
            details_window = ObjectDetailsWindow(obj_type, self.current_dump, self.comparison_dump)
            details_window.show()
            details_window.activateWindow()
            self.details_windows.append(details_window)
            self.logger.info(f"Opened details window for object type {obj_type}")
        except Exception as e:
            self.logger.exception("Failed to open object details window")
            show_error_message(self, "Details Error", f"Failed to open object details: {str(e)}", get_traceback())

    def show_column_menu(self, position):
        """
        Show a context menu to toggle column visibility.

        Args:
            position: The position where the menu is requested.
        """
        menu = QMenu()
        model = self.table_view.model()
        if not model:
            return

        header = self.table_view.horizontalHeader()
        for col in range(model.columnCount()):
            action = QAction(model.headerData(col, Qt.Orientation.Horizontal), self)
            action.setCheckable(True)
            action.setChecked(not self.table_view.isColumnHidden(col))
            action.toggled.connect(lambda checked, c=col: self.table_view.setColumnHidden(c, not checked))
            menu.addAction(action)

        menu.exec(self.table_view.viewport().mapToGlobal(position))

    def apply_filters(self):
        """
        Apply the current filters and search criteria to the table.
        """
        try:
            if not self.current_dump:
                self.show_message("Info", "Load a heap dump first to apply filters")
                return

            # Get filter values
            min_size_text = self.min_size_input.text().strip()
            max_size_text = self.max_size_input.text().strip()
            min_size = int(min_size_text) if min_size_text else None
            max_size = int(max_size_text) if max_size_text else None

            # Get selected object type
            selected_type = self.type_filter_combo.currentText()
            if selected_type == "All Types":
                selected_type = None

            # Get status filters
            status_filter = []
            if self.comparison_dump:
                for status, checkbox in self.status_checkboxes.items():
                    if checkbox.isChecked():
                        status_filter.append(status)

            # Get search values
            search_id = self.search_id_input.text().strip()
            search_attr = self.search_attr_input.text().strip()

            # Apply filters through controller
            self.controller.handle_apply_filters(min_size, max_size, selected_type, status_filter, search_id, search_attr)
        except Exception as e:
            self.logger.exception("Failed to apply filters")
            show_error_message(self, "Filter Error", f"Failed to apply filters: {str(e)}", get_traceback())

    def reset_filters(self):
        """
        Reset all filters and search fields to their default values.
        """
        self.min_size_input.clear()
        self.max_size_input.clear()
        self.type_filter_combo.setCurrentIndex(0)
        for checkbox in self.status_checkboxes.values():
            checkbox.setChecked(False)
        self.search_id_input.clear()
        self.search_attr_input.clear()
        
        # Refresh table with no filters
        if self.current_dump:
            if self.comparison_dump:
                self.populate_table(self.current_dump, comparison=True)
            else:
                self.populate_table(self.current_dump)

    def update_type_filter_combo(self, object_types):
        """
        Update the object type filter combo box with available types.
        
        Args:
            object_types (list): List of object type names.
        """
        current_selection = self.type_filter_combo.currentText()
        self.type_filter_combo.clear()
        self.type_filter_combo.addItem("All Types")
        for obj_type in sorted(object_types):
            self.type_filter_combo.addItem(obj_type)
        
        # Restore selection if possible
        index = self.type_filter_combo.findText(current_selection)
        if index >= 0:
            self.type_filter_combo.setCurrentIndex(index)

    def export_results(self, file_path):
        """
        Export current table data to a file.

        Args:
            file_path (str): Path to the export file.
        """
        try:
            if file_path.endswith('.csv'):
                self._export_to_csv(file_path)
            elif file_path.endswith('.json'):
                self._export_to_json(file_path)
            elif file_path.endswith('.xlsx'):
                self._export_to_excel(file_path)
            else:
                raise ValueError("Unsupported export format. Use .csv, .json, or .xlsx")

            self.logger.info(f"Results exported to {file_path}")
        except Exception as e:
            self.logger.exception("Failed to export results")
            raise

    def _export_to_csv(self, file_path):
        """
        Export data to CSV format.

        Args:
            file_path (str): Path to the CSV file.
        """
        model = self.table_view.model()
        if not model:
            raise ValueError("No data to export")

        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)

            # Write headers
            headers = []
            for col in range(model.columnCount()):
                headers.append(model.headerData(col, Qt.Orientation.Horizontal))
            writer.writerow(headers)

            # Write data
            for row in range(model.rowCount()):
                row_data = []
                for col in range(model.columnCount()):
                    item = model.index(row, col).data()
                    row_data.append(item if item is not None else '')
                writer.writerow(row_data)

    def _export_to_json(self, file_path):
        """
        Export data to JSON format.

        Args:
            file_path (str): Path to the JSON file.
        """
        model = self.table_view.model()
        if not model:
            raise ValueError("No data to export")

        data = []

        # Write headers
        headers = []
        for col in range(model.columnCount()):
            headers.append(model.headerData(col, Qt.Orientation.Horizontal))

        # Write data
        for row in range(model.rowCount()):
            row_data = {}
            for col in range(model.columnCount()):
                item = model.index(row, col).data()
                row_data[headers[col]] = item if item is not None else ''
            data.append(row_data)

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    def _export_to_excel(self, file_path):
        """
        Export data to Excel format.

        Args:
            file_path (str): Path to the Excel file.
        """
        model = self.table_view.model()
        if not model:
            raise ValueError("No data to export")

        wb = Workbook()
        ws = wb.active
        
        # Write headers with styling
        headers = []
        for col in range(model.columnCount()):
            header = model.headerData(col, Qt.Orientation.Horizontal)
            headers.append(header)
            c = ws.cell(row=1, column=col + 1, value=header)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center', vertical='center')

        # Write data
        for row in range(model.rowCount()):
            for col in range(model.columnCount()):
                item = model.index(row, col).data()
                value = item if item is not None else ''
                ws.cell(row=row + 2, column=col + 1, value=value)

        # Adjust column widths
        for col_idx in range(len(headers)):
            col_letter = chr(65 + col_idx) if col_idx < 26 else 'A' + chr(65 + col_idx - 26)
            max_length = len(str(headers[col_idx]))
            for row_idx in range(2, model.rowCount() + 2):
                c = ws.cell(row=row_idx, column=col_idx + 1)
                cell_value = c.value if hasattr(c, 'value') and c.value else ''
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 2, 50)
            try:
                ws.column_dimensions[col_letter].width = adjusted_width
            except:
                pass

        # Freeze header row
        try:
            ws.freeze_panes = 'A2'
        except:
            pass

        wb.save(file_path)
